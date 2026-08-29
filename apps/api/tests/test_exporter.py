from openpyxl import load_workbook

from app.domain import NeedProfile
from app.exporter import export_plan
from app.features.builds.planner import generate_plans


def test_export_contains_summary_parts_and_compatibility_sheets(tmp_path):
    plan = generate_plans(NeedProfile())[1]
    output = export_plan(plan, tmp_path, "export-test")
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["方案概览", "配件明细", "兼容性检查"]
    assert workbook["配件明细"].max_row == len(plan.items) + 1

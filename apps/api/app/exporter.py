from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.domain import BuildPlan


def export_plan(plan: BuildPlan, output_dir: Path, job_id: str) -> str:
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"build-plan-{job_id}.xlsx"
    workbook = Workbook()
    summary = workbook.active
    if summary is None:  # openpyxl 的类型定义允许 None，正常新建工作簿不会出现该情况。
        raise RuntimeError("无法创建方案概览工作表")
    summary.title = "方案概览"
    summary.append(["智能装机方案", plan.title])
    summary.append(["预算", plan.budget])
    summary.append(["总价参考", plan.total_price])
    summary.append(["预计功耗", f"{plan.estimated_power_w}W"])
    summary.append(["性能参考分", plan.performance_score])
    summary.append(["说明", plan.summary])
    summary["A1"].font = Font(bold=True, color="FFFFFF")
    summary["B1"].font = Font(bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="0D9488")
    summary["B1"].fill = PatternFill("solid", fgColor="0D9488")

    details = workbook.create_sheet("配件明细")
    details.append(["配置项", "名称", "品牌", "参考价", "来源", "锁定", "选择理由"])
    for cell in details[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0D9488")
    for item in plan.items:
        details.append(
            [
                item.slot.value,
                item.part.name,
                item.part.brand,
                item.part.price,
                item.part.source,
                "是" if item.locked else "否",
                item.reason,
            ]
        )

    issues = workbook.create_sheet("兼容性检查")
    issues.append(["级别", "标题", "详情", "关联配置项"])
    for issue in plan.compatibility:
        issues.append([issue.severity, issue.title, issue.detail, ", ".join(issue.related_slots)])
    workbook.save(path)
    return str(path)
